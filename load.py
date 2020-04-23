from openpyxl import load_workbook
from findFlightSize import findSize


# def that returns an array of the 5th column in the active sheet from the excel document
def loadData():
    doc = load_workbook(filename='CPH_KEA sep2019-1.xlsx')
    ws = doc.active
    array = []

    for col in ws.iter_rows(min_col=5, max_col=5, min_row=2, values_only=True):
        for value in col:
            array.append(findSize(str(value)))

    return array
