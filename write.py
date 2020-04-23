from openpyxl import load_workbook
from load import loadData

doc = load_workbook(filename='CPH_KEA sep2019-1.xlsx')
# extracting the active sheet from the document
ws = doc.active

# loadData() returns a array of all the flighnumbers
result = loadData()

columns = 0
# loop through collumn 7 in ac
for col in ws.iter_rows(min_col=7, max_col=7, min_row=2):
    for cell in col:
        cell.value = result[columns]
        columns = columns + 1

doc.save('CPH_flykoder_opdateret.xlsx')